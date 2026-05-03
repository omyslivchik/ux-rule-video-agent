"""
Шаг 4. Строит итоговый отчёт из output/chunks/scene_analyses.jsonl.

- Читает все записи из scene_analyses.jsonl
- Отбирает сцены, связанные с созданием правила
- Один запрос к OpenRouter → получает Markdown-отчёт
- Парсит из Markdown таблицы (CJM, Decision Map, Боли)
- Сохраняет:
  - output/reports/final_report.md
  - output/reports/ux_rule_creation_analysis.xlsx   (листы: Summary, CJM, Decision Map, Pains)
"""

import json
import os
import re
import sys
import time
from pathlib import Path

import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from tqdm import tqdm

SKIP_STAGES = {"не относится к созданию правила"}


def _is_relevant(chunk: dict) -> bool:
    """Сцена релевантна если: либо stage не в SKIP_STAGES,
    либо в ней есть реальный контент (decision, risk_or_pain, short_summary)."""
    stage = chunk.get("rule_creation_stage", "")
    if stage not in SKIP_STAGES:
        return True
    has_content = any([
        chunk.get("decision", "").strip(),
        chunk.get("risk_or_pain", "").strip(),
        chunk.get("short_summary", "").strip(),
    ])
    return has_content


# ---------------------------------------------------------------------------
# Запрос к OpenRouter (текст)
# ---------------------------------------------------------------------------

def _call_openrouter_text(prompt: str, env: dict, config: dict) -> str:
    headers = {
        "Authorization": f"Bearer {env['OPENROUTER_API_KEY']}",
        "Content-Type": "application/json",
        "HTTP-Referer": env.get("OPENROUTER_HTTP_REFERER", ""),
        "X-Title": env.get("OPENROUTER_X_TITLE", ""),
    }
    payload = {
        "model": env["OPENROUTER_MODEL"],
        "messages": [{"role": "user", "content": prompt}],
        "temperature": config["openrouter_temperature"],
        "max_tokens": config["openrouter_max_tokens"],
    }

    max_retries: int = config["max_retries"]
    backoff: float = config["retry_backoff_sec"]
    last_err = None

    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.post(
                env["OPENROUTER_BASE_URL"],
                headers=headers,
                json=payload,
                timeout=config["request_timeout_sec"],
            )
            resp.raise_for_status()
            return resp.json()["choices"][0]["message"]["content"]
        except (requests.RequestException, KeyError) as e:
            last_err = e
            wait = backoff * (2 ** (attempt - 1))
            print(
                f"[build_report] Ошибка попытка {attempt}/{max_retries}: {e}. Жду {wait:.1f}с...",
                file=sys.stderr,
            )
            time.sleep(wait)

    raise RuntimeError(f"[build_report] Не удалось получить отчёт: {last_err}")


# ---------------------------------------------------------------------------
# Парсинг Markdown-таблиц
# ---------------------------------------------------------------------------

def _parse_md_table(md_text: str, section_keyword: str) -> pd.DataFrame:
    """
    Находит первую Markdown-таблицу после строки, содержащей section_keyword.
    """
    lines = md_text.splitlines()
    start = None
    for i, line in enumerate(lines):
        if section_keyword.lower() in line.lower():
            for j in range(i + 1, len(lines)):
                stripped = lines[j].strip()
                if stripped.startswith("|"):
                    start = j
                    break
            if start is not None:
                break

    if start is None:
        return pd.DataFrame()

    table_lines = []
    for line in lines[start:]:
        if not line.strip().startswith("|"):
            break
        table_lines.append(line.strip())

    if len(table_lines) < 2:
        return pd.DataFrame()

    def split_row(row: str) -> list[str]:
        return [c.strip() for c in row.strip("|").split("|")]

    headers = split_row(table_lines[0])
    rows = []
    for line in table_lines[2:]:
        if re.match(r"^\|[\s\-:|]+\|$", line):
            continue
        rows.append(split_row(line))

    return pd.DataFrame(rows, columns=headers) if rows else pd.DataFrame(columns=headers)


def _extract_summary_text(md_text: str) -> str:
    """Извлекает текст раздела '1. Краткое резюме'."""
    match = re.search(
        r"##\s*1\.\s*Краткое резюме\s*(.*?)(?=##\s*2\.|$)",
        md_text,
        re.DOTALL | re.IGNORECASE,
    )
    return match.group(1).strip() if match else md_text[:500]


# ---------------------------------------------------------------------------
# Запись Excel
# ---------------------------------------------------------------------------

def _parse_timecode_ms(timecode_str: str):
    """Парсит таймкод в миллисекунды. Форматы: MM:SS, HH:MM:SS, 0:40:34–..."""
    # берём только первый таймкод если диапазон ('0:00–28:00' или '00:40:34–00:41:00')
    tc = timecode_str.split("–")[0].split("-")[0].strip()
    parts = re.findall(r"\d+", tc)
    if len(parts) == 3:
        # HH:MM:SS
        h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
        return (h * 3600 + m * 60 + s) * 1000
    elif len(parts) == 2:
        # MM:SS
        m, s = int(parts[0]), int(parts[1])
        return (m * 60 + s) * 1000
    return None


def _find_closest_frame(timecode_str: str, session_dir: Path):
    """По таймкоду находит ближайший кадр в frames_kept. Форматы: MM:SS, HH:MM:SS."""
    frames_dir = session_dir / "frames_kept"
    if not frames_dir.exists():
        return None
    frame_files = sorted(frames_dir.glob("frame_*.jpg"))
    if not frame_files:
        return None

    target_ms = _parse_timecode_ms(timecode_str)
    if target_ms is None:
        return frame_files[0]

    best, best_diff = frame_files[0], float("inf")
    for fpath in frame_files:
        try:
            ts_ms = int(fpath.stem.replace("frame_", ""))
        except ValueError:
            continue
        diff = abs(ts_ms - target_ms)
        if diff < best_diff:
            best_diff, best = diff, fpath
    return best



def _add_screenshots_to_sheet(wb, sheet_name: str, timecode_col_name: str, session_dir: Path) -> None:
    """Добавляет колонку Скриншот в произвольный лист, подбирая кадр по таймкоду."""
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    img_w, img_h = 320, 180
    row_height_pt = img_h * 0.75

    timecode_col_idx = None
    for cell in ws[1]:
        if cell.value and timecode_col_name.lower() in str(cell.value).lower():
            timecode_col_idx = cell.column
            break

    screenshot_col = ws.max_column + 1
    ws.cell(row=1, column=screenshot_col, value="Скриншот")
    ws.column_dimensions[get_column_letter(screenshot_col)].width = 46

    for row_idx in range(2, ws.max_row + 1):
        timecode_str = ""
        if timecode_col_idx:
            cell_val = ws.cell(row=row_idx, column=timecode_col_idx).value
            timecode_str = str(cell_val) if cell_val else ""
        frame_path = _find_closest_frame(timecode_str, session_dir)
        if frame_path is None:
            continue
        try:
            img = XLImage(str(frame_path))
            img.width = img_w
            img.height = img_h
            ws.add_image(img, f"{get_column_letter(screenshot_col)}{row_idx}")
            ws.row_dimensions[row_idx].height = row_height_pt
        except Exception as e:
            ws.cell(row=row_idx, column=screenshot_col, value=f"[ошибка: {e}]")


def _write_excel(report_md: str, chunks: list[dict], out_path: Path, session_dir: Path = None) -> None:
    slepok_df = _parse_md_table(report_md, "Слепок действий")
    cjm_df = _parse_md_table(report_md, "CJM")
    decision_df = _parse_md_table(report_md, "Decision Map")
    if decision_df.empty:
        decision_df = _parse_md_table(report_md, "Decision map")
    pains_df = _parse_md_table(report_md, "Наблюдаемые боли")
    if pains_df.empty:
        pains_df = _parse_md_table(report_md, "Pains")

    summary_text = _extract_summary_text(report_md)
    summary_df = pd.DataFrame([{
        "Параметр": "Краткое резюме",
        "Значение": summary_text,
    }, {
        "Параметр": "Всего сцен проанализировано",
        "Значение": len(chunks),
    }, {
        "Параметр": "Сцен с ошибками",
        "Значение": sum(1 for c in chunks if "error" in c),
    }])

    with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        if not slepok_df.empty:
            slepok_df.to_excel(writer, sheet_name="Слепок действий", index=False)
        else:
            pd.DataFrame(columns=["Нет данных"]).to_excel(writer, sheet_name="Слепок действий", index=False)
        if not cjm_df.empty:
            cjm_df.to_excel(writer, sheet_name="CJM", index=False)
        else:
            pd.DataFrame(columns=["Нет данных"]).to_excel(writer, sheet_name="CJM", index=False)
        if not decision_df.empty:
            decision_df.to_excel(writer, sheet_name="Decision Map", index=False)
        else:
            pd.DataFrame(columns=["Нет данных"]).to_excel(writer, sheet_name="Decision Map", index=False)
        if not pains_df.empty:
            pains_df.to_excel(writer, sheet_name="Pains", index=False)
        else:
            pd.DataFrame(columns=["Нет данных"]).to_excel(writer, sheet_name="Pains", index=False)

        if session_dir is not None:
            if not slepok_df.empty:
                _add_screenshots_to_sheet(writer.book, "Слепок действий", "таймкод", session_dir)
            if not cjm_df.empty:
                _add_screenshots_to_sheet(writer.book, "CJM", "таймкод", session_dir)

    print(f"[build_report] Excel сохранён → {out_path}")


# ---------------------------------------------------------------------------
# Точка входа
# ---------------------------------------------------------------------------

def run(config: dict, project_root: Path, session_name: str = "final", session_dir: Path = None) -> None:
    load_dotenv(project_root / ".env")

    env = {
        "OPENROUTER_API_KEY": os.environ.get("OPENROUTER_API_KEY", ""),
        "OPENROUTER_MODEL": os.environ.get("OPENROUTER_MODEL", ""),
        "OPENROUTER_BASE_URL": os.environ.get("OPENROUTER_BASE_URL", ""),
        "OPENROUTER_HTTP_REFERER": os.environ.get("OPENROUTER_HTTP_REFERER", ""),
        "OPENROUTER_X_TITLE": os.environ.get("OPENROUTER_X_TITLE", ""),
    }

    if not env["OPENROUTER_API_KEY"] or env["OPENROUTER_API_KEY"].startswith("your_"):
        print("[build_report] OPENROUTER_API_KEY не задан в .env", file=sys.stderr)
        sys.exit(1)

    if session_dir is None:
        session_dir = project_root / "output"
    analyses_path = session_dir / "chunks" / "scene_analyses.jsonl"
    if not analyses_path.exists():
        print(f"[build_report] Не найден {analyses_path}. Сначала запустите analyze_scenes.", file=sys.stderr)
        sys.exit(1)

    all_chunks = [
        json.loads(line)
        for line in analyses_path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]

    relevant_chunks = [c for c in all_chunks if _is_relevant(c)]

    skipped = len(all_chunks) - len(relevant_chunks)
    print(f"[build_report] Всего сцен: {len(all_chunks)}, используется: {len(relevant_chunks)}, пропущено: {skipped}")

    if not relevant_chunks:
        print("[build_report] Нет релевантных сцен для отчёта.", file=sys.stderr)
        sys.exit(1)

    prompt_template = (project_root / "prompts" / "final_report_prompt.md").read_text(encoding="utf-8")
    chunks_json = json.dumps(relevant_chunks, ensure_ascii=False, indent=2)
    final_prompt = f"{prompt_template}\n\n---\n\n## Данные сцен:\n\n```json\n{chunks_json}\n```"

    print("[build_report] Отправляю запрос в OpenRouter...")
    report_md = _call_openrouter_text(final_prompt, env, config)

    session_dir.mkdir(parents=True, exist_ok=True)

    md_path = session_dir / f"{session_name}_report.md"
    md_path.write_text(report_md, encoding="utf-8")
    print(f"[build_report] Markdown → {md_path}")

    xlsx_path = session_dir / f"{session_name}_analysis.xlsx"
    _write_excel(report_md, all_chunks, xlsx_path, session_dir=session_dir)


if __name__ == "__main__":
    root = Path(__file__).resolve().parent.parent
    cfg = json.loads((root / "config.json").read_text(encoding="utf-8"))
    run(cfg, root)
